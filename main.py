import datetime
import random
import os
import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError

def save_to_excel(file_name, data_rows):
    sheet_name = "総合"
    headers = ["日付", "ユーザー名", "フォロー数", "フォロワー数", "ポスト数"]
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
    else:
        wb = openpyxl.Workbook()
        wb.active.title = sheet_name
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        ws.append(headers)
    for row in data_rows:
        ws.append(row)
    wb.save(file_name)

def scrape():
    input_csv = "targets.csv"
    output_xlsx = "x_reports.xlsx"
    
    if not os.path.exists(input_csv):
        # テスト用にファイルがなければ作成
        with open(input_csv, "w") as f: f.write("elonmusk\nnasa")
    
    with open(input_csv, 'r') as f:
        usernames = [line.strip() for line in f if line.strip()]

    results = []
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with sync_playwright() as p:
        # GitHub上では headless=True 必須
        browser = p.chromium.launch(headless=True)
        # 人間らしく見せるための設定
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        page = context.new_page()

        for username in usernames:
            print(f"🔍 Checking: @{username}")
            current_data = {"followers": None, "following": None, "posts": None}

            def handle_response(response):
                if "UserByScreenName" in response.url and response.status == 200:
                    try:
                        data = response.json()
                        user = data['data']['user']['result']['legacy']
                        current_data["followers"] = user['followers_count']
                        current_data["following"] = user['friends_count']
                        current_data["posts"] = user['statuses_count']
                    except: pass

            page.on("response", handle_response)
            try:
                page.goto(f"https://x.com/{username}", wait_until="domcontentloaded", timeout=30000)
                page.wait_for_timeout(5000) # 通信待ち
                if current_data["followers"] is not None:
                    results.append([now_str, username, current_data["following"], current_data["followers"], current_data["posts"]])
                    print(f" ✅ Success: {username}")
                else:
                    print(f" ❌ Failed (No Data): {username}")
            except Exception as e:
                print(f" ⚠️ Error: {e}")
            page.remove_listener("response", handle_response)
        browser.close()

    if results:
        save_to_excel(output_xlsx, results)

if __name__ == "__main__":
    scrape()